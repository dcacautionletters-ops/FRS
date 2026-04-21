import streamlit as st
import pandas as pd
import io
import plotly.express as px
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- 1. UI CONFIGURATION ---
st.set_page_config(page_title="VMS Universal Reporting", layout="wide")
MASTER_PASSWORD = "VMS@123"

st.markdown("""
    <style>
    .stApp {
        background: linear-gradient(rgba(15, 23, 42, 0.92), rgba(15, 23, 42, 0.92)), 
                    url("https://images.unsplash.com/photo-1451187580459-43490279c0fa?ixlib=rb-1.2.1&auto=format&fit=crop&w=1950&q=80");
        background-size: cover; background-attachment: fixed;
    }
    .welcome-note { 
        background: linear-gradient(to right, #00d2ff, #92fe9d); 
        -webkit-background-clip: text; -webkit-text-fill-color: transparent; 
        font-size: 48px !important; font-weight: 700; text-align: center; margin: 40px 0 10px 0;
    }
    .glass-metric {
        background: rgba(255, 255, 255, 0.03); backdrop-filter: blur(15px);
        border: 1px solid rgba(255, 255, 255, 0.1); border-radius: 15px;
        padding: 25px; margin: 10px 0; text-align: center;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
        position: relative;
    }
    .metric-value { font-size: 42px; font-weight: 800; color: #92fe9d; }
    .metric-title { color: #ffffff; font-size: 14px; font-weight: 600; text-transform: uppercase; }
    
    /* Layer 1: Surgical Button Styling to make headers interactive */
    .stButton > button {
        background: transparent !important;
        border: none !important;
        padding: 0 !important;
        color: inherit !important;
        width: 100%;
    }
    </style>
    """, unsafe_allow_html=True)

# --- 2. AUTHENTICATION & STATE ---
if 'authenticated' not in st.session_state: st.session_state.authenticated = False
# New State Layers
if 'sel_section' not in st.session_state: st.session_state.sel_section = None
if 'sel_subject' not in st.session_state: st.session_state.sel_subject = None

if not st.session_state.authenticated:
    st.markdown('<p class="welcome-note">VMS Reporting System</p>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.4, 1])
    with col2:
        p = st.text_input("Password", type="password")
        if st.button("Access Dashboard", use_container_width=True):
            if p == MASTER_PASSWORD: st.session_state.authenticated = True; st.rerun()
            else: st.error("Access Denied")
    st.stop()

# --- 3. CORE LOGIC (KEEPING YOUR EXACT FUNCTIONS) ---
KEYWORDS_TO_IGNORE = ["BADMINTON", "BASKETBALL", "CROSS FITNESS", "SWIMMING", "ZUMBA", "TABLE TENNIS", 
                      "FREESLOT", "FREE SLOT", "SOFT SKILL", "ATOM", "DSA"]
ATT_COL_NAME = "Attended Hours with Approved Leave Percentage"

def is_valid_subject(subject_name):
    s_upper = str(subject_name).upper()
    return not any(bad in s_upper for bad in KEYWORDS_TO_IGNORE)

def get_bracket_summary(data_df, cols, subjects, threshold):
    summary_data = []
    for sub in subjects:
        sub_vals = pd.to_numeric(data_df[data_df[cols['subject']] == sub][cols['attendance']], errors='coerce').dropna().round(2)
        b1 = len(sub_vals[(sub_vals >= 0) & (sub_vals < 50)])
        b2 = len(sub_vals[(sub_vals >= 50) & (sub_vals < 60)])
        b3a = len(sub_vals[(sub_vals >= 60) & (sub_vals < 64.5)])
        b3b = len(sub_vals[(sub_vals >= 64.5) & (sub_vals < 70)])
        b4 = len(sub_vals[(sub_vals >= 70) & (sub_vals < 75)])
        row = {"Subject": sub}
        total = 0
        if threshold > 0: row["0.00-49.99"] = b1; total += b1
        if threshold > 50: row["50.00-59.99"] = b2; total += b2
        if threshold > 60:
            row["60.00-64.49"] = b3a
            row["64.50-69.99"] = b3b 
            total += (b3a + b3b)
        if threshold > 70: row["70.00-74.99"] = b4; total += b4
        row["Total"] = total
        summary_data.append(row)
    return pd.DataFrame(summary_data)

def apply_styles(ws, threshold, is_summary=False):
    thin = Side(style='thin', color="4D4D4D")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    crit_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid") 
    warn_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") 
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font, cell.fill, cell.border = Font(bold=True, color="FFFFFF"), h_fill, border
        ws.column_dimensions[cell.column_letter].width = 20
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border, cell.alignment = border, Alignment(horizontal="center")
            if not is_summary and cell.column > 5:
                try:
                    val = float(cell.value)
                    if val < 70: cell.fill, cell.font = crit_fill, Font(bold=True, color="FFFFFF")
                    elif 70 <= val < threshold: cell.fill, cell.font = warn_fill, Font(bold=True, color="000000")
                except: pass

def process_grid(data_df, cols, batch_subjects, low_thresh, high_thresh, show_all=False):
    if data_df.empty: return None, None
    data_df = data_df.copy()
    data_df[cols['attendance']] = pd.to_numeric(data_df[cols['attendance']], errors='coerce').round(2)
    full_grid = data_df.pivot_table(index=[cols['roll'], cols['name'], cols['batch'], cols['sem']],
                                    columns=cols['subject'], values=cols['attendance'], sort=False).reset_index()
    final_subjects = [s for s in batch_subjects if is_valid_subject(s)]
    for sub in final_subjects:
        if sub not in full_grid.columns: full_grid[sub] = None
        full_grid[sub] = pd.to_numeric(full_grid[sub], errors='coerce').round(2)
    theory_cols = [c for c in final_subjects if not any(x in str(c).upper() for x in ["LAB", "PRACTICAL", "WORKSHOP"])]
    full_grid['Theory Avg'] = full_grid[theory_cols].mean(axis=1).round(2)
    full_grid['Final Avg'] = full_grid[final_subjects].mean(axis=1).round(2)
    grid_mask = (full_grid[final_subjects] >= low_thresh) & (full_grid[final_subjects] <= high_thresh)
    shortage_grid = full_grid if show_all else full_grid[grid_mask.any(axis=1)].copy()
    if shortage_grid.empty: return None, None
    active_mask = (shortage_grid[final_subjects] >= low_thresh) & (shortage_grid[final_subjects] <= high_thresh)
    shortage_grid['Subjects in Range'] = active_mask.sum(axis=1)
    sub_counts = active_mask.sum()
    if not show_all:
        for sub in final_subjects:
            shortage_grid[sub] = shortage_grid[sub].apply(lambda x: x if (pd.notnull(x) and low_thresh <= x <= high_thresh) else "")
    shortage_grid.insert(0, 'Sl No.', range(1, len(shortage_grid) + 1))
    return shortage_grid, sub_counts

# --- 4. DASHBOARD INTERFACE ---
uploaded_file = st.file_uploader("📂 Upload Universal Attendance File", type=["xlsx"])

if uploaded_file:
    df_raw = pd.read_excel(uploaded_file, header=None).head(15)
    h_row = 0
    for i, row in df_raw.iterrows():
        if any("ROLL NO" in str(x).upper() for x in row.values):
            h_row = i; break
    df = pd.read_excel(uploaded_file, header=h_row)
    c_map = {'sem': df.columns[5]} 
    for c in df.columns:
        cs = str(c).strip()
        if "Roll No" in cs: c_map['roll'] = c
        elif "Student Name" in cs: c_map['name'] = c
        elif "Batch" in cs: c_map['batch'] = c
        elif any(x in cs for x in ["Course", "Subject"]): c_map['subject'] = c
        elif ATT_COL_NAME in cs: c_map['attendance'] = c

    df = df[df[c_map['subject']].apply(is_valid_subject)]
    df['Dept'] = df[c_map['batch']].astype(str).apply(lambda x: x.split()[0].upper())
    
    with st.sidebar:
        st.markdown("### 🛠️ Global Parameters")
        c_l, c_h = st.columns(2)
        with c_l: low_v = st.number_input("From (%)", 0.00, 100.00, 0.00, 0.01, format="%.2f")
        with c_h: high_v = st.number_input("To (%)", 0.00, 100.00, 75.00, 0.01, format="%.2f")
        dept_choice = st.selectbox("Select Department", ["All Departments"] + sorted(df['Dept'].unique()))
        exclude_subs = st.multiselect("Exclude Subjects", sorted(df[c_map['subject']].unique()))
        if st.button("Clear Selection Filters"): 
            st.session_state.sel_section = None
            st.session_state.sel_subject = None
            st.rerun()
        if st.button("Logout"): st.session_state.authenticated = False; st.rerun()

    if exclude_subs: df = df[~df[c_map['subject']].isin(exclude_subs)]
    active_depts = [dept_choice] if dept_choice != "All Departments" else sorted(df['Dept'].unique())

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summaries, subject_impact = [], pd.Series(dtype=float)
        # Store data for dynamic filtering
        all_section_data = {}

        tabs = st.tabs(["📊 COMMAND CENTER"] + [f"💎 {d}" for d in active_depts])

        for d_idx, dept in enumerate(active_depts):
            d_df = df[df['Dept'] == dept]
            series_list = sorted(list(set(f"{b.split()[0]} {next((p for p in b.split() if p.isdigit()), 'Series')}" for b in d_df[c_map['batch']].astype(str).unique())))
            
            with tabs[d_idx+1]:
                for series in series_list:
                    s_df = d_df[d_df[c_map['batch']].astype(str).str.contains(series.split()[0]) & d_df[c_map['batch']].astype(str).str.contains(series.split()[-1])]
                    s_subs = sorted([s for s in s_df[c_map['subject']].unique() if is_valid_subject(s)])
                    
                    for sec in sorted(s_df[c_map['batch']].unique()):
                        sec_df = s_df[s_df[c_map['batch']] == sec]
                        grid, counts = process_grid(sec_df, c_map, s_subs, low_v, high_v)
                        if grid is not None:
                            all_section_data[sec] = {'grid': grid, 'counts': counts, 'series': series}
                            summaries.append({'Section': sec, 'Count': len(grid)-1, 'Series': series})
                            subject_impact = subject_impact.add(counts, fill_value=0)

        with tabs[0]:
            if summaries:
                sum_df = pd.DataFrame(summaries)
                
                # --- Layer 2: Subject Glass Headers (Rows by Batch Year) ---
                st.markdown("### 📚 Subject Focus (Click to Filter)")
                batches = sorted(df[c_map['batch']].astype(str).unique())
                for b_name in batches:
                    b_subs = sorted(df[df[c_map['batch']] == b_name][c_map['subject']].unique())
                    st.write(f"**{b_name} Subjects:**")
                    cols = st.columns(min(len(b_subs), 6))
                    for i, sub in enumerate(b_subs):
                        with cols[i % 6]:
                            if st.button(f"{sub}", key=f"sub_{b_name}_{sub}"):
                                st.session_state.sel_subject = sub
                                st.session_state.sel_section = None # Reset section when subject clicked
                
                # --- Layer 1: Interactive Section Glass Headers ---
                st.markdown("### 🏢 Section Overview (Click to Filter)")
                m_cols = st.columns(min(len(sum_df), 4))
                for idx, row in sum_df.iterrows():
                    with m_cols[idx % 4]:
                        # Wrap glass-metric in a button
                        if st.button("", key=f"glass_{row['Section']}"):
                            st.session_state.sel_section = row['Section']
                            st.session_state.sel_subject = None # Reset subject when section clicked
                        st.markdown(f'''
                            <div class="glass-metric" style="margin-top: -85px; pointer-events: none;">
                                <div class="metric-title">{row["Section"]}</div>
                                <div class="metric-value">{row["Count"]}</div>
                            </div>
                        ''', unsafe_allow_html=True)

                # --- Dynamic Filtering Logic ---
                filtered_sum_df = sum_df.copy()
                filtered_impact = subject_impact.copy()
                
                if st.session_state.sel_section:
                    st.subheader(f"📍 Filtering by Section: {st.session_state.sel_section}")
                    filtered_sum_df = sum_df[sum_df['Section'] == st.session_state.sel_section]
                    display_grid = all_section_data[st.session_state.sel_section]['grid']
                    st.dataframe(display_grid, hide_index=True)
                    
                elif st.session_state.sel_subject:
                    st.subheader(f"📖 Filtering by Subject: {st.session_state.sel_subject}")
                    # Re-calculating impact and grid for only this subject
                    sub_list = []
                    for s_key, s_val in all_section_data.items():
                        g = s_val['grid']
                        if st.session_state.sel_subject in g.columns:
                            sub_only = g[g[st.session_state.sel_subject] != ""]
                            if not sub_only.empty: sub_list.append(sub_only)
                    
                    if sub_list:
                        final_sub_df = pd.concat(sub_list)
                        st.dataframe(final_sub_df, hide_index=True)
                
                # Charts update dynamically based on filtered data
                c1, c2 = st.columns(2)
                with c1: 
                    st.plotly_chart(px.bar(filtered_sum_df, x='Section', y='Count', color='Section',
                                     title="Dynamic Section Distribution", template="plotly_dark"), use_container_width=True)
                with c2:
                    impact_df = filtered_impact.reset_index()
                    impact_df.columns = ['Subject', 'Students']
                    if st.session_state.sel_subject: impact_df = impact_df[impact_df['Subject'] == st.session_state.sel_subject]
                    st.plotly_chart(px.pie(impact_df[impact_df['Students']>0], names='Subject', values='Students', hole=0.4, 
                                     title="Dynamic Subject Impact", template="plotly_dark"), use_container_width=True)
                
                # Sheet Export Logic
                sum_df.to_excel(writer, sheet_name='SUMMARY', index=False)
                for s_key, s_val in all_section_data.items():
                    sn_sec = str(s_key).replace("/", "-")[:31]
                    s_val['grid'].to_excel(writer, sheet_name=sn_sec, index=False)
                    apply_styles(writer.sheets[sn_sec], high_v)
                    
            else: st.info("No data in current range.")

    st.download_button(f"📥 Download Report", output.getvalue(), "VMS_Report.xlsx", use_container_width=True)
